"""Helper module for file operations, data processing, and S3 interactions."""

import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import lru_cache
from typing import List
import warnings
from math import log2

import requests
import geopandas as gpd
import pandas as pd
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import boto3
from boto3.s3.transfer import TransferConfig
from dotenv import load_dotenv

load_dotenv()


def write_parquet(filepath, df=None):
    """Write dataframe to Parquet format.

    Args:
        filepath (str): Base path for the output file
        df (pd.DataFrame): DataFrame to write
    """
    df.to_parquet(f"{filepath}.parquet", index=False, compression="brotli")
    print(f"Wrote Parquet: {filepath}")


def write_csv(filepath, df=None):
    """Write dataframe to CSV format.

    Args:
        filepath (str): Base path for the output file
        df (pd.DataFrame): DataFrame to write
    """
    df.to_csv(f"{filepath}.csv", index=False)
    print(f"Wrote CSV: {filepath}")


def write_csv_parquet(filepath, df=None):
    """Write dataframe to both CSV and Parquet formats.

    Args:
        filepath (str): Base path for the output files
        df (pd.DataFrame): DataFrame to write
    """
    df.to_csv(f"{filepath}.csv", index=False)
    df.to_parquet(f"{filepath}.parquet", index=False, compression="brotli")
    print(f"Wrote CSV + Parquet: {filepath}")


def write_csv_parquet_excel(filepath, df=None):
    """Write dataframe to CSV, Parquet, and a formatted Excel file.

    Args:
        filepath (str): Base path for the output files
        df (pd.DataFrame): DataFrame to write
    """
    df.to_csv(f"{filepath}.csv", index=False)
    df.to_parquet(f"{filepath}.parquet", index=False, compression="brotli")

    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.freeze_panes = "A2"

    header_fill = PatternFill(start_color="9B1C1C", end_color="9B1C1C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Aptos")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws.row_dimensions[1].height = 20

    dtypes = {col: df[col].dtype for col in df.columns}

    for row_idx, row_vals in enumerate(df.itertuples(index=False, name=None), 2):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, (col_name, value) in enumerate(zip(df.columns, row_vals), 1):
            # Convert pandas NA/NaT/NaN to None so openpyxl renders blank
            if value is pd.NaT or (isinstance(value, float) and pd.isna(value)):
                value = None
            # Convert pandas Timestamp to Python datetime
            elif hasattr(value, "to_pydatetime"):
                value = value.to_pydatetime()
            # Convert numpy int/float to Python native
            elif hasattr(value, "item"):
                value = value.item()

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Aptos")
            dtype = dtypes[col_name]

            if pd.api.types.is_datetime64_any_dtype(dtype):
                cell.number_format = "YYYY-MM-DD"
                cell.alignment = center
            elif pd.api.types.is_integer_dtype(dtype):
                cell.number_format = "#,##0"
                cell.alignment = right
            elif pd.api.types.is_float_dtype(dtype):
                cell.number_format = "#,##0.##"
                cell.alignment = right
            else:
                cell.alignment = left

    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for val in df[col_name]:
            if val is not None and not (isinstance(val, float) and pd.isna(val)):
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max_len + 4

    wb.save(f"{filepath}.xlsx")
    print(f"Wrote CSV + Parquet + Excel: {filepath}")


@lru_cache(maxsize=1)
def get_s3_client():
    """Create and cache S3 client."""
    return boto3.client(
        "s3",
        aws_access_key_id=os.getenv("S3_ACCESS_KEY_ID"),
        aws_secret_access_key=os.getenv("S3_SECRET_ACCESS_KEY"),
    )


@lru_cache(maxsize=1)
def get_r2_client():
    """Create and cache R2 client."""
    return boto3.client(
        "s3",
        endpoint_url=f"https://{os.getenv('R2_ACCOUNT_ID')}.r2.cloudflarestorage.com",
        aws_access_key_id=os.getenv("R2_ACCESS_KEY_ID"),
        aws_secret_access_key=os.getenv("R2_SECRET_ACCESS_KEY"),
        region_name="auto",
    )


def get_transfer_config():
    """Get optimized transfer configuration."""
    return TransferConfig(
        multipart_threshold=1024 * 25,
        max_concurrency=10,
        multipart_chunksize=1024 * 25,
        use_threads=True,
    )


def upload_single(
    client=None, bucket=None, source_file=None, cloud_file=None, content_type="application/json"
):
    """Upload a single file to S3/R2.

    Args:
        client: S3/R2 client
        bucket: S3/R2 bucket
        source_file: Source file path
        cloud_file: Cloud file path
        content_type: Content type of the file
    """
    try:
        time_start = time.time()
        client.upload_file(
            source_file,
            bucket,
            cloud_file,
            Config=get_transfer_config(),
            ExtraArgs={"ContentType": content_type},
        )
        duration = f"{time.time() - time_start:.1f} seconds"
        return source_file, True, f"SUCCESS ({duration}): {bucket}/{cloud_file}"
    except Exception as e:
        return source_file, False, f"FAILURE: {bucket}/{source_file}\n\n{e}"


def upload_bulk(
    client=None, bucket=None, files_to_upload=None, max_workers=50, content_type="application/json"
):
    """Upload multiple files to S3/R2 in parallel.

    Args:
        client: S3/R2 client
        bucket: S3/R2 bucket
        files_to_upload: List of tuples (source_file, cloud_file)
        max_workers: Maximum number of workers
        content_type: Content type of the files
    """
    results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_file = {
            executor.submit(upload_single, client, bucket, source, cloud, content_type): source
            for source, cloud in files_to_upload
        }
        for future in as_completed(future_to_file):
            source_file, success, message = future.result()
            results[source_file] = (success, message)
            print(message)

    return [(source, msg) for source, (success, msg) in results.items() if not success]


def purge_cf_cache(file_keys: list[str], base_url: str) -> None:
    """
    Purge Cloudflare cache for a list of file keys.

    Args:
        file_keys: List of R2 object keys e.g. ["parties/dropdown.json", "parties/all.json"]
        base_url: Public base URL e.g. "https://internal.electiondata.my"

    Requires env vars:
        CF_ZONE_ID: Cloudflare Zone ID
        CF_CACHE_PURGE_TOKEN: Cloudflare API token with Cache Purge permission
    """

    zone_id = os.getenv("CF_ZONE_ID")
    api_token = os.getenv("CF_CACHE_PURGE_TOKEN")

    urls = [f"{base_url.rstrip('/')}/{key}" for key in file_keys]

    # Cloudflare allows max 30 URLs per request
    for i in range(0, len(urls), 30):
        batch = urls[i : i + 30]
        resp = requests.post(
            f"https://api.cloudflare.com/client/v4/zones/{zone_id}/purge_cache",
            headers={"Authorization": f"Bearer {api_token}"},
            json={"files": batch},
        )
        resp.raise_for_status()
        result = resp.json()
        if not result.get("success"):
            raise RuntimeError(f"Cache purge failed: {result.get('errors')}")
        plural = "s" if len(batch) > 1 else ""
        print(f"Purged {len(batch)} URL{plural} from cache")


def copy_bulk_within_r2(
    client, source_bucket, dest_bucket, prefix, dest_prefix=None, max_workers=50, exclude=None
):
    """Copy all objects under a prefix from one R2 bucket to another in parallel.

    Args:
        client: R2 client
        source_bucket: Source bucket name
        dest_bucket: Destination bucket name
        prefix: Key prefix to filter source objects e.g. "results/"
        dest_prefix: Key prefix for destination keys — defaults to prefix if not set
        max_workers: Maximum number of parallel copy workers
        exclude: List of keys to skip e.g. ["elections/all.json"]
    """
    if dest_prefix is None:
        dest_prefix = prefix
    exclude_set = set(exclude or [])

    paginator = client.get_paginator("list_objects_v2")
    keys = [
        obj["Key"]
        for page in paginator.paginate(Bucket=source_bucket, Prefix=prefix)
        for obj in page.get("Contents", [])
        if obj["Key"] not in exclude_set
    ]
    print(
        f"\nCopying {len(keys):,.0f} files from {source_bucket}/{prefix} to {dest_bucket}/{dest_prefix}"
    )

    def copy_object(key):
        dest_key = dest_prefix + key[len(prefix) :]
        client.copy_object(
            CopySource={"Bucket": source_bucket, "Key": key},
            Bucket=dest_bucket,
            Key=dest_key,
        )
        return key

    total = len(keys)
    milestone = max(1, total // 100)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(copy_object, key): key for key in keys}
        for i, future in enumerate(as_completed(futures), start=1):
            future.result()
            if i % milestone == 0 or i == total:
                print(f"  {i / total:.0%} ({i:,.0f}/{total:,.0f})")

    print(f"Done — copied {total:,.0f} files to {dest_bucket}")


def purge_cf_cache_prefix(prefixes: list[str]) -> None:
    """Purge Cloudflare cache by URL prefix (Pro plan+). Max 30 prefixes per call.

    Args:
        prefixes: List of URL prefixes without scheme e.g. ["internal.electiondata.my/og-image/"]

    Requires env vars:
        CF_ZONE_ID: Cloudflare Zone ID
        CF_CACHE_PURGE_TOKEN: Cloudflare API token with Cache Purge permission
    """
    zone_id = os.getenv("CF_ZONE_ID")
    api_token = os.getenv("CF_CACHE_PURGE_TOKEN")

    for i in range(0, len(prefixes), 30):
        batch = prefixes[i : i + 30]
        resp = requests.post(
            f"https://api.cloudflare.com/client/v4/zones/{zone_id}/purge_cache",
            headers={"Authorization": f"Bearer {api_token}"},
            json={"prefixes": batch},
        )
        resp.raise_for_status()
        result = resp.json()
        if not result.get("success"):
            raise RuntimeError(f"Cache purge failed: {result.get('errors')}")
        plural = "es" if len(batch) > 1 else ""
        print(f"Purged {len(batch)} prefix{plural} from cache")


def generate_slug(x):
    """Generate URL-friendly slug from string.

    Args:
        x (str): Input string

    Returns:
        str: URL-friendly slug
    """
    slug = re.sub(r"[^a-zA-Z0-9\s]", "", x.replace("-", " "))
    slug = slug.replace(" ", "-").lower()
    return slug


def get_states(my: int = 0, codes: int = 0) -> List[str]:
    """Get list of Malaysian states.

    Args:
        my (int): Whether to include Malaysia (country)
        code (int): Whether to return full name (0), text code (1), or integer code (2)
    Returns:
        List[str]: List of states as full name, text codes, or integer codes
    """
    data = {
        "Malaysia": ["MYS", 0],
        "Perlis": ["PLS", 1],
        "Kedah": ["KDH", 2],
        "Kelantan": ["KTN", 3],
        "Terengganu": ["TRG", 4],
        "Pulau Pinang": ["PNG", 5],
        "Perak": ["PRK", 6],
        "Pahang": ["PHG", 7],
        "Selangor": ["SGR", 8],
        "W.P. Kuala Lumpur": ["KUL", 9],
        "W.P. Putrajaya": ["PJY", 10],
        "Negeri Sembilan": ["NSN", 11],
        "Melaka": ["MLK", 12],
        "Johor": ["JHR", 13],
        "W.P. Labuan": ["LBN", 14],
        "Sabah": ["SBH", 15],
        "Sarawak": ["SWK", 16],
    }
    state_names = list(data.keys())
    state_codes = [data[x][0] for x in state_names]
    state_order = [data[x][1] for x in state_names]
    if codes == 0:
        return state_names[1 - my :]
    if codes == 1:
        return state_codes[1 - my :]
    if codes == 2:
        return state_order[1 - my :]
    raise ValueError("Invalid code type")


def capitalize_sentence(sentence):
    """Capitalize first word and title case remaining words in sentence.

    Args:
        sentence (str): Input sentence

    Returns:
        str: Properly capitalized sentence
    """
    words = sentence.split()
    return " ".join([words[0].upper()] + [word.title() for word in words[1:]])


def compute_summary(
    df,
    election=None,
    exclude_state=None,
    include_state=None,
    col_groupby=["coalition", "party"],
    col_seat="seat",
):
    """
    Compute summary statistics for a given election and (optionally) selected/excluded states.

    Args:
        df (pandas.DataFrame): DataFrame containing election data. Must have columns: 'election', 'state', 'votes', and all in col_groupby.
        election (str, optional): Election identifier string. If provided, filters rows to only this election.
        include_state (list of str, optional): List of state names to include. If provided, retains only these states.
        exclude_state (list of str, optional): List of state names to exclude. Ignored if null OR if include_state is specified.
        col_groupby (list of str, optional): Columns to group by. Default is ['coalition', 'party'].

    Returns:
        pandas.DataFrame: DataFrame with columns: col_groupby, 'votes' (comma formatted), and 'votes_perc' (rounded % of total votes per group).
    """
    if election:
        df = df[df.election == election]
    if include_state:
        df = df[df.state.isin(include_state)]
    elif exclude_state:
        df = df[~df.state.isin(exclude_state)]
    df["votes_perc"] = df["votes"] / df["votes"].sum() * 100
    df["seats"] = 0

    wf = df.sort_values(by="votes", ascending=True).drop_duplicates(subset=col_seat, keep="last")
    df.loc[df.index.isin(wf.index), "seats"] = 1

    rf = (
        df[col_groupby + ["votes", "votes_perc", "seats"]]
        .groupby(col_groupby)
        .sum()
        .sort_values(by=[col_groupby[0], "votes"], ascending=False)
    )
    rf.votes_perc = rf.votes_perc.round(1)
    rf["votes"] = rf["votes"].apply(lambda x: f"{x:,}")
    return rf


def get_final_cols(file_type="ballots"):
    """Get final columns for a given file type. Extra layer of validation that all cols are present.

    Args:
        file_type (str): Type of file ("ballots" or "stats")

    Returns:
        list: List of columns
    """
    if file_type == "ballots":
        return [
            "date",
            "election",
            "state",
            "seat",
            "ballot_order",
            "candidate_uid",
            "name_on_ballot",
            "name",
            "sex",
            "ethnicity",
            "age",
            "party_on_ballot",
            "party_uid",
            "party",
            "coalition_uid",
            "coalition",
            "votes",
            "votes_perc",
            "rank",
            "result",
        ]
    if file_type == "stats":
        return [
            "date",
            "election",
            "state",
            "seat",
            "voters_total",
            "ballots_issued",
            "ballots_not_returned",
            "votes_rejected",
            "votes_valid",
            "majority",
            "n_candidates",
            "voter_turnout",
            "majority_perc",
            "votes_rejected_perc",
            "ballots_not_returned_perc",
        ]
    return []


def calculate_zoom(feature):
    """
    Calculate an appropriate zoom level for a GeoPandas GeoSeries/GeoDataFrame feature.

    Args:
        feature (GeoSeries or GeoDataFrame): A GeoSeries or GeoDataFrame whose total_bounds are used.

    Returns:
        float: A zoom level value, bounded between 0 and 20, with a bias for "zooming out" above level 3.
    """
    minx, miny, maxx, maxy = feature.total_bounds
    x_margin = (maxx - minx) * 0.1
    y_margin = (maxy - miny) * 0.1
    minx -= x_margin
    maxx += x_margin
    miny -= y_margin
    maxy += y_margin

    lat_span = maxy - miny
    lon_span = maxx - minx
    scale = max(lon_span / 360, lat_span / 180)
    zoom = -log2(scale) + 1

    zoom = min(max(zoom, 0), 20)
    return zoom - 2 if zoom > 3 else zoom


def row_zoom(geom):
    """
    Calculate the zoom level for an individual geometry object.

    Args:
        geom (shapely Geometry): The geometry for which to calculate zoom.

    Returns:
        float: The calculated zoom level for the geometry.
    """
    return calculate_zoom(gpd.GeoSeries([geom]))


def get_center_and_zoom(files=[]):
    """
    Compute zoom levels and center coordinates for a list of GeoParquet files.

    Args:
        files (list, optional): List of file paths to GeoParquet files. Defaults to empty list.

    Returns:
        pandas.DataFrame: DataFrame with columns ['seat', 'state', 'center', 'zoom'], containing
            one row for each seat with its corresponding calculated zoom and center.
            The center is [lon, lat] rounded to 6 decimal places and zoom is rounded to 1 decimal place.
    """
    res = pd.DataFrame(columns=["seat", "state", "center", "zoom"])
    for f in files:
        gf = gpd.read_parquet(f)
        gf = gf.to_crs(epsg=4236)
        seat_type = "dun" if "dun" in f else "parlimen"

        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Geometry is in a geographic CRS. Results from 'centroid' are likely incorrect.*",
            )
            gf["center"] = gf.centroid
        gf["center"] = gf["center"].apply(lambda p: [round(p.x, 6), round(p.y, 6)])
        gf["zoom"] = gf["geometry"].apply(row_zoom)
        gf.zoom = gf.zoom.round(1)
        gf = gf[[seat_type, "state", "center", "zoom"]].rename(columns={seat_type: "seat"})
        res = gf.copy() if len(res) == 0 else pd.concat([res, gf], axis=0, ignore_index=True)
    return res


def get_voter_pyramid_from_vr(vr=None, reference_year=None):
    """
    Generate a voter age-sex pyramid dataset from a voter registry Parquet file.

    Args:
        vr (str, optional): Path to the voter registry Parquet file.
        reference_year (int): Year against which birth_year is compared to compute age.

    Returns:
        pandas.DataFrame: DataFrame with columns:
            - 'slug': unique identifier for each seat-state combination,
            - 'sex': gender of voters,
            - 'age': age of voters (18 to 100),
            - 'voters': number of voters for each age, sex, and seat.

    The function reads the Parquet file, computes per seat, sex, and single-year age group
    the total voter counts, normalizing state + seat as a slug.
    """
    query = f"""
    WITH base AS (
        SELECT state, parlimen AS seat, sex, LEAST({reference_year} - birth_year, 100) AS age
        FROM read_parquet('{vr}')
        UNION ALL
        SELECT state, dun AS seat, sex, LEAST({reference_year} - birth_year, 100) AS age
        FROM read_parquet('{vr}')
        WHERE NOT state LIKE 'W.P.%'
    ),
    counts AS (
        SELECT state, seat, sex, age, COUNT(*) AS voters
        FROM base
        WHERE age >= 18
        GROUP BY state, seat, sex, age
    ),
    seats AS (
        SELECT DISTINCT state, parlimen AS seat FROM read_parquet('{vr}')
        UNION ALL
        SELECT DISTINCT state, dun AS seat FROM read_parquet('{vr}')
        WHERE NOT state LIKE 'W.P.%'
    ),
    ages AS (
        SELECT unnest(generate_series(18, 100)) AS age
    ),
    sexes AS (
        SELECT DISTINCT sex FROM read_parquet('{vr}')
    ),
    grid AS (
        SELECT s.state, s.seat, sx.sex, a.age
        FROM seats s
        CROSS JOIN sexes sx
        CROSS JOIN ages a
    )
    SELECT
        g.state,
        g.seat,
        g.sex,
        g.age::INTEGER AS age,
        COALESCE(c.voters, 0) AS voters
    FROM grid g
    LEFT JOIN counts c USING (state, seat, sex, age)
    ORDER BY g.seat, g.sex, g.age
    """

    df = duckdb.query(query).fetchdf()
    df.seat = df.seat + ", " + df.state
    df = df.rename(columns={"seat": "slug"})
    df = df.drop(columns=["state"])
    df.slug = df.slug.apply(generate_slug)
    return df


def get_age_x_ethnicity_from_vr(vr=None, reference_year=None):
    """
    Returns a DataFrame aggregating counts of voters by age and ethnicity for each seat and state
    from the given voter registry parquet file.

    It:
    - Normalizes certain Sabah and Sarawak ethnic groups for aggregation.
    - Counts registered voters per (state, seat, ethnicity, age band).
    - Provides cross-joined enumeration of all observed state-ethnicity-seat trios.
    - Only includes voters aged 18 and above.

    Inputs:
    - vr: Path to .parquet file containing voter registry data.
    - reference_year: Year against which birth_year is compared to compute age.

    Outputs:
    - DataFrame with voting age population counts grouped by state, seat, ethnicity, and age band.
    """
    query = f"""
    WITH base AS (
        SELECT
            state,
            parlimen AS seat,
            CASE
                WHEN state = 'Sabah'   AND ethnicity IN ('Kadazan', 'Dusun')          THEN 'Kadazan-Dusun'
                WHEN state = 'Sabah'   AND ethnicity IN ('Bumi Sarawak', 'Indian')    THEN 'Other'
                WHEN state = 'Sarawak' AND ethnicity IN ('Malay', 'Melanau')          THEN 'Malay-Melanau'
                WHEN state = 'Sarawak' AND ethnicity IN ('Bumi Sabah', 'Indian')      THEN 'Other'
                ELSE ethnicity
            END AS ethnicity,
            LEAST({reference_year} - birth_year, 100) AS age
        FROM read_parquet('{vr}')
        UNION ALL
        SELECT
            state,
            dun AS seat,
            CASE
                WHEN state = 'Sabah'   AND ethnicity IN ('Kadazan', 'Dusun')          THEN 'Kadazan-Dusun'
                WHEN state = 'Sabah'   AND ethnicity IN ('Bumi Sarawak', 'Indian')    THEN 'Other'
                WHEN state = 'Sarawak' AND ethnicity IN ('Malay', 'Melanau')          THEN 'Malay-Melanau'
                WHEN state = 'Sarawak' AND ethnicity IN ('Bumi Sabah', 'Indian')      THEN 'Other'
                ELSE ethnicity
            END AS ethnicity,
            LEAST({reference_year} - birth_year, 100) AS age
        FROM read_parquet('{vr}')
        WHERE NOT state LIKE 'W.P.%'
    ),
    state_ethnicities AS (
        SELECT DISTINCT state, ethnicity
        FROM base
        WHERE age >= 18
    ),
    seats AS (
        SELECT DISTINCT state, parlimen AS seat FROM read_parquet('{vr}')
        UNION ALL
        SELECT DISTINCT state, dun AS seat FROM read_parquet('{vr}')
        WHERE NOT state LIKE 'W.P.%'
    ),
    grid AS (
        SELECT s.state, s.seat, e.ethnicity
        FROM seats s
        JOIN state_ethnicities e ON s.state = e.state
    ),
    counts AS (
        SELECT
            state, seat, ethnicity,
            COUNT(*)                                        AS "all-ages",
            COUNT(*) FILTER (WHERE age BETWEEN 18 AND 29)  AS "18-29",
            COUNT(*) FILTER (WHERE age BETWEEN 30 AND 39)  AS "30-39",
            COUNT(*) FILTER (WHERE age BETWEEN 40 AND 49)  AS "40-49",
            COUNT(*) FILTER (WHERE age BETWEEN 50 AND 59)  AS "50-59",
            COUNT(*) FILTER (WHERE age BETWEEN 60 AND 69)  AS "60-69",
            COUNT(*) FILTER (WHERE age BETWEEN 70 AND 79)  AS "70-79",
            COUNT(*) FILTER (WHERE age >= 80)               AS "80+"
        FROM base
        WHERE age >= 18
        GROUP BY state, seat, ethnicity
    ),
    ethnic_rows AS (
        SELECT
            g.state, g.seat, g.ethnicity,
            COALESCE(c."all-ages", 0) AS "all-ages",
            COALESCE(c."18-29",  0) AS "18-29",
            COALESCE(c."30-39",  0) AS "30-39",
            COALESCE(c."40-49",  0) AS "40-49",
            COALESCE(c."50-59",  0) AS "50-59",
            COALESCE(c."60-69",  0) AS "60-69",
            COALESCE(c."70-79",  0) AS "70-79",
            COALESCE(c."80+",    0) AS "80+"
        FROM grid g
        LEFT JOIN counts c USING (state, seat, ethnicity)
    ),
    overall_rows AS (
        SELECT
            state, seat, 'all-ethnicities' AS ethnicity,
            SUM("all-ages") AS "all-ages",
            SUM("18-29")  AS "18-29",
            SUM("30-39")  AS "30-39",
            SUM("40-49")  AS "40-49",
            SUM("50-59")  AS "50-59",
            SUM("60-69")  AS "60-69",
            SUM("70-79")  AS "70-79",
            SUM("80+")    AS "80+"
        FROM ethnic_rows
        GROUP BY state, seat
    )
    SELECT * FROM (
        SELECT * FROM overall_rows
        UNION ALL
        SELECT * FROM ethnic_rows
    ) t
    ORDER BY seat, CASE WHEN ethnicity = 'all-ethnicities' THEN 0 ELSE 1 END, "all-ages" DESC
    """
    df = duckdb.query(query).fetchdf()
    for c in df.columns[3:]:
        df[c] = df[c].astype(int)
    df.seat = df.seat + ", " + df.state
    df = df.rename(columns={"seat": "slug"})
    df = df.drop(columns=["state"])
    df.ethnicity = df.ethnicity.apply(generate_slug)
    df.slug = df.slug.apply(generate_slug)
    return df
